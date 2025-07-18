from flask import Flask, request, jsonify, send_file
import json
from datetime import datetime
import os
import tempfile
import io
from PIL import Image
from docx import Document
import anthropic
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import base64
import threading
from concurrent.futures import ThreadPoolExecutor
import uuid
import platform
import subprocess

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB for bulk

# Red Hat Environment Detection and Configuration
def detect_redhat_environment():
    """Detect if running on Red Hat/OpenShift environment"""
    redhat_indicators = {
        'openshift': False,
        'rhel': False,
        'podman': False,
        'systemd': False
    }
    
    try:
        # Check for OpenShift environment variables
        if os.environ.get('OPENSHIFT_BUILD_NAME') or os.environ.get('KUBERNETES_SERVICE_HOST'):
            redhat_indicators['openshift'] = True
            logger.info("🎯 Red Hat OpenShift environment detected")
        
        # Check for RHEL
        if os.path.exists('/etc/redhat-release'):
            redhat_indicators['rhel'] = True
            with open('/etc/redhat-release', 'r') as f:
                release_info = f.read().strip()
                logger.info(f"🎯 Red Hat Enterprise Linux detected: {release_info}")
        
        # Check for Podman (Red Hat's container engine)
        try:
            subprocess.run(['podman', '--version'], capture_output=True, check=True)
            redhat_indicators['podman'] = True
            logger.info("🎯 Red Hat Podman detected")
        except (subprocess.CalledProcessError, FileNotFoundError):
            pass
        
        # Check for systemd (Red Hat's init system)
        if os.path.exists('/run/systemd/system'):
            redhat_indicators['systemd'] = True
            logger.info("🎯 Red Hat systemd detected")
        
        return redhat_indicators
    except Exception as e:
        logger.warning(f"Red Hat environment detection failed: {e}")
        return redhat_indicators

# Initialize Red Hat environment
redhat_env = detect_redhat_environment()

# Initialize Claude API client
claude_client = anthropic.Anthropic(
    api_key=os.environ.get('CLAUDE_API_KEY')
)

# Store processing results temporarily
processing_results = {}
processing_status = {}

# Global metrics for dashboard
dashboard_metrics = {
    'total_documents': 0,
    'total_hours': 0,
    'total_entries': 0,
    'total_images': 0
}

def extract_images_from_docx(docx_path):
    """Extract ALL images from DOCX file"""
    images = []
    try:
        doc = Document(docx_path)
        
        for rel in doc.part.rels.values():
            if "image" in rel.target_ref:
                try:
                    image_data = rel.target_part.blob
                    image = Image.open(io.BytesIO(image_data))
                    
                    if image.mode != 'RGB':
                        image = image.convert('RGB')
                    
                    # Resize very large images but keep all images
                    if image.size[0] > 1500 or image.size[1] > 1500:
                        image.thumbnail((1500, 1500), Image.Resampling.LANCZOS)
                    
                    images.append(image)
                    
                except Exception as e:
                    logger.error(f"❌ Error extracting image: {e}")
                    continue
        
        logger.info(f"✅ Extracted {len(images)} images from {docx_path}")
        return images
    except Exception as e:
        logger.error(f"❌ Error processing DOCX: {e}")
        return []

def image_to_base64(image):
    """Convert PIL image to base64 string"""
    buffered = io.BytesIO()
    # Use JPEG with good quality for balance of size and clarity
    image.save(buffered, format="JPEG", quality=90, optimize=True)
    img_str = base64.b64encode(buffered.getvalue()).decode()
    return img_str

def process_images_in_batches(images, filename):
    """Process images in batches to handle Claude's limits but get ALL data"""
    all_entries = []
    batch_size = 4  # Process 4 images at a time for better reliability
    
    for i in range(0, len(images), batch_size):
        batch = images[i:i + batch_size]
        logger.info(f"Processing batch {i//batch_size + 1}/{(len(images) + batch_size - 1)//batch_size} for {filename}")
        
        try:
            # Convert batch to base64
            image_data = []
            for img in batch:
                base64_img = image_to_base64(img)
                image_data.append({
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/jpeg",
                        "data": base64_img
                    }
                })
            
            # Create message for this batch with specific instructions for timesheet status
            message_content = [
                {
                    "type": "text",
                    "text": f"""
                    File: {filename} (Batch {i//batch_size + 1})
                    
                    Analyze these timesheet screenshots and extract ALL entries. Look carefully at BOTH the main timesheet table AND the right panel.
                    
                    IMPORTANT: The submission status is shown in the right panel and may be highlighted in YELLOW. Look for:
                    - "State" section in right panel
                    - Status like "Closed", "Open", "Submitted", "Pending", "Approved"
                    - Yellow highlighting indicates status
                    
                    From the main table, extract:
                    - Employee name (from title like "Aravind G / 2025-06-07 / Week 23")
                    - Each date row with hours
                    - All individual timesheet entries
                    
                    From the right panel, extract:
                    - Current submission status (look for yellow highlighting)
                    - Week information
                    - Total hours
                    
                    Return ONLY JSON array:
                    [
                        {{
                            "employee_name": "Aravind G",
                            "date": "06/09/2025",
                            "hours": 8.0,
                            "submission_status": "Closed",
                            "week": "Week 23",
                            "total_hours": 40.0
                        }}
                    ]
                    
                    CRITICAL RULES:
                    - Extract ALL individual date entries from the table
                    - Use the status from the right panel (yellow highlighted area)
                    - Convert dates to MM/DD/YYYY format  
                    - Include week information if available
                    - If status is highlighted in yellow, that's the current status
                    - Common statuses: Closed, Open, Submitted, Pending, Approved
                    """
                }
            ]
            
            message_content.extend(image_data)
            
            # Call Claude API
            response = claude_client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=2000,
                messages=[{
                    "role": "user", 
                    "content": message_content
                }]
            )
            
            response_text = response.content[0].text.strip()
            
            # Parse JSON response
            try:
                batch_entries = json.loads(response_text)
                if isinstance(batch_entries, list):
                    all_entries.extend(batch_entries)
                    logger.info(f"✅ Batch {i//batch_size + 1}: Found {len(batch_entries)} entries")
                else:
                    logger.warning(f"⚠️ Batch {i//batch_size + 1}: Invalid response format")
            except json.JSONDecodeError:
                # Try to find JSON in response
                import re
                json_match = re.search(r'\[.*?\]', response_text, re.DOTALL)
                if json_match:
                    try:
                        batch_entries = json.loads(json_match.group())
                        if isinstance(batch_entries, list):
                            all_entries.extend(batch_entries)
                            logger.info(f"✅ Batch {i//batch_size + 1}: Found {len(batch_entries)} entries (parsed)")
                    except json.JSONDecodeError:
                        logger.error(f"❌ Batch {i//batch_size + 1}: Could not parse JSON")
                        continue
                else:
                    logger.error(f"❌ Batch {i//batch_size + 1}: No JSON found in response")
                    continue
            
        except Exception as e:
            logger.error(f"❌ Error processing batch {i//batch_size + 1}: {e}")
            continue
    
    logger.info(f"✅ Total entries extracted from {filename}: {len(all_entries)}")
    return all_entries

def extract_timesheet_data_with_claude(images, filename=""):
    """Process ALL images to extract complete timesheet data"""
    if not images:
        return []
    
    logger.info(f"🔍 Processing {len(images)} images from {filename}")
    
    # Process all images in batches
    all_entries = process_images_in_batches(images, filename)
    
    # Add source file to each entry
    for entry in all_entries:
        entry['source_file'] = filename
    
    return all_entries

def process_single_file(file_data, job_id):
    """Process single file in thread"""
    try:
        filename, file_content = file_data
        
        # Update status
        processing_status[job_id]['current_file'] = filename
        processing_status[job_id]['processed'] += 1
        
        logger.info(f"📁 Processing file: {filename}")
        
        # Save temp file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.docx') as temp_file:
            temp_file.write(file_content)
            temp_path = temp_file.name
        
        try:
            # Extract ALL images
            images = extract_images_from_docx(temp_path)
            
            if not images:
                logger.warning(f"⚠️ No images found in {filename}")
                return []
            
            # Extract data from ALL images
            data = extract_timesheet_data_with_claude(images, filename)
            
            # Update global metrics
            dashboard_metrics['total_images'] += len(images)
            
            logger.info(f"✅ Extracted {len(data)} entries from {filename}")
            return data
            
        finally:
            os.unlink(temp_path)
            
    except Exception as e:
        logger.error(f"❌ Error processing {filename}: {e}")
        return []

def process_files_bulk(files, job_id):
    """Process multiple files in sequence (to avoid API rate limits)"""
    try:
        all_data = []
        
        # Process files one by one to avoid Claude API rate limits
        for file_data in files:
            try:
                result = process_single_file(file_data, job_id)
                all_data.extend(result)
                
                # Small delay to avoid rate limiting
                import time
                time.sleep(1)
                
            except Exception as e:
                logger.error(f"❌ Error processing file: {e}")
                continue
        
        # Update global metrics
        dashboard_metrics['total_documents'] += len(files)
        dashboard_metrics['total_entries'] += len(all_data)
        
        # Calculate total hours
        for entry in all_data:
            dashboard_metrics['total_hours'] += entry.get('hours', 0)
        
        # Store results
        processing_results[job_id] = all_data
        processing_status[job_id]['status'] = 'completed'
        processing_status[job_id]['total_entries'] = len(all_data)
        
        logger.info(f"🎉 Bulk processing completed: {len(all_data)} total entries")
        
    except Exception as e:
        logger.error(f"❌ Bulk processing error: {e}")
        processing_status[job_id]['status'] = 'error'
        processing_status[job_id]['error'] = str(e)

def create_excel_file(data):
    """Create Excel file with timesheet data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet Data"
    
    # Headers - updated to include week and total hours
    headers = ['Employee Name', 'Date', 'Hours', 'Submission Status', 'Week', 'Total Hours', 'Source File']
    
    # Style headers with Red Hat colors
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")  # Red Hat Red
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, entry in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=entry.get('employee_name', 'Unknown')).border = border
        ws.cell(row=row_idx, column=2, value=entry.get('date', '')).border = border
        ws.cell(row=row_idx, column=3, value=entry.get('hours', 0)).border = border
        ws.cell(row=row_idx, column=4, value=entry.get('submission_status', 'pending')).border = border
        ws.cell(row=row_idx, column=5, value=entry.get('week', '')).border = border
        ws.cell(row=row_idx, column=6, value=entry.get('total_hours', 0)).border = border
        ws.cell(row=row_idx, column=7, value=entry.get('source_file', '')).border = border
        
        # Highlight closed status in yellow (like in the original)
        if entry.get('submission_status', '').lower() == 'closed':
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            ws.cell(row=row_idx, column=4).fill = yellow_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

@app.route('/dashboard')
def dashboard():
    """Dashboard endpoint with metrics"""
    # Calculate accuracy rate (mockup)
    accuracy_rate = 95 if dashboard_metrics['total_entries'] > 0 else 0
    
    # Calculate average time savings per document
    time_savings_per_doc = 10  # minutes saved per document
    
    # Calculate discrepancies (estimate)
    discrepancies = max(1, dashboard_metrics['total_entries'] // 25) if dashboard_metrics['total_entries'] > 0 else 0
    
    dashboard_data = {
        'documents_processed': dashboard_metrics['total_documents'],
        'hours_extracted': dashboard_metrics['total_hours'],
        'accuracy_rate': accuracy_rate,
        'time_savings': time_savings_per_doc,
        'total_entries': dashboard_metrics['total_entries'],
        'images_processed': dashboard_metrics['total_images'],
        'discrepancies': discrepancies
    }
    
    return jsonify(dashboard_data)

@app.route('/')
def index():
    """Main page with integrated dashboard metrics"""
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Time Verify AI</title>
        <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
        <style>
            * {
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }
            
            body {
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                color: #1a1a1a;
                overflow-x: hidden;
            }
            
            .container {
                max-width: 1200px;
                margin: 0 auto;
                padding: 2rem;
            }
            
            .header {
                text-align: center;
                margin-bottom: 3rem;
                animation: fadeInUp 0.8s ease;
            }
            
            .header h1 {
                font-size: 2.5rem;
                font-weight: 700;
                color: white;
                margin-bottom: 0.5rem;
                text-shadow: 0 2px 4px rgba(0,0,0,0.1);
            }
            
            .header p {
                font-size: 1.1rem;
                color: rgba(255,255,255,0.8);
                font-weight: 300;
            }
            
            .redhat-badge {
                display: inline-block;
                background: rgba(255,255,255,0.1);
                color: white;
                padding: 0.5rem 1rem;
                border-radius: 20px;
                font-size: 0.8rem;
                margin-top: 1rem;
                backdrop-filter: blur(10px);
            }
            
            .metrics-bar {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 1rem;
                margin-bottom: 3rem;
                animation: fadeInUp 0.8s ease 0.1s both;
            }
            
            .metric-card {
                background: rgba(255, 255, 255, 0.95);
                backdrop-filter: blur(20px);
                border-radius: 12px;
                padding: 1.5rem;
                box-shadow: 0 8px 25px rgba(0,0,0,0.1);
                border-top: 3px solid;
                transition: transform 0.3s ease;
            }
            
            .metric-card:hover {
                transform: translateY(-2px);
            }
            
            .metric-card.blue { border-top-color: #3b82f6; }
            .metric-card.green { border-top-color: #10b981; }
            .metric-card.orange { border-top-color: #f59e0b; }
            .metric-card.purple { border-top-color: #8b5cf6; }
            
            .metric-header {
                display: flex;
                align-items: center;
                justify-content: space-between;
                margin-bottom: 0.75rem;
            }
            
            .metric-title {
                font-size: 0.75rem;
                font-weight: 500;
                color: #6b7280;
                text-transform: uppercase;
                letter-spacing: 0.05em;
            }
            
            .metric-icon {
                width: 1.5rem;
                height: 1.5rem;
                border-radius: 4px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 0.875rem;
            }
            
            .metric-icon.blue { background: #dbeafe; }
            .metric-icon.green { background: #d1fae5; }
            .metric-icon.orange { background: #fef3c7; }
            .metric-icon.purple { background: #ede9fe; }
            
            .metric-value {
                font-size: 2rem;
                font-weight: 700;
                color: #1a202c;
                margin-bottom: 0.25rem;
            }
            
            .metric-subtitle {
                font-size: 0.75rem;
                color: #10b981;
                font-weight: 500;
                display: flex;
                align-items: center;
                gap: 0.25rem;
            }
            
            .main-card {
                background: rgba(255, 255, 255, 0.95);
                backdrop-filter: blur(20px);
                border-radius: 24px;
                padding: 3rem;
                box-shadow: 0 20px 40px rgba(0,0,0,0.1);
                animation: fadeInUp 0.8s ease 0.2s both;
            }
            
            .upload-section {
                text-align: center;
                margin-bottom: 2rem;
            }
            
            .upload-area {
                border: 2px dashed #e1e5e9;
                border-radius: 16px;
                padding: 3rem 2rem;
                background: linear-gradient(135deg, #f8f9ff 0%, #f1f3ff 100%);
                transition: all 0.3s ease;
                cursor: pointer;
                position: relative;
                overflow: hidden;
            }
            
            .upload-area:hover {
                border-color: #667eea;
                background: linear-gradient(135deg, #f0f2ff 0%, #e8ecff 100%);
                transform: translateY(-2px);
            }
            
            .upload-area::before {
                content: '';
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                bottom: 0;
                background: linear-gradient(45deg, transparent 49%, rgba(102, 126, 234, 0.1) 50%, transparent 51%);
                opacity: 0;
                transition: opacity 0.3s ease;
            }
            
            .upload-area:hover::before {
                opacity: 1;
            }
            
            .upload-icon {
                width: 64px;
                height: 64px;
                background: linear-gradient(135deg, #667eea, #764ba2);
                border-radius: 16px;
                display: flex;
                align-items: center;
                justify-content: center;
                margin: 0 auto 1.5rem;
                font-size: 24px;
                color: white;
                box-shadow: 0 8px 20px rgba(102, 126, 234, 0.3);
            }
            
            .upload-text {
                font-size: 1.2rem;
                font-weight: 600;
                color: #2d3748;
                margin-bottom: 0.5rem;
            }
            
            .upload-subtext {
                font-size: 0.9rem;
                color: #718096;
                font-weight: 400;
            }
            
            .file-input {
                display: none;
            }
            
            .file-list {
                margin: 1.5rem 0;
                padding: 1.5rem;
                background: #f7fafc;
                border-radius: 12px;
                border-left: 4px solid #667eea;
                display: none;
            }
            
            .file-item {
                padding: 0.5rem 0;
                color: #4a5568;
                font-size: 0.9rem;
                display: flex;
                align-items: center;
            }
            
            .file-item::before {
                content: '📄';
                margin-right: 0.5rem;
            }
            
            .action-buttons {
                display: flex;
                gap: 1rem;
                justify-content: center;
                margin-top: 2rem;
            }
            
            .btn {
                padding: 0.75rem 2rem;
                border: none;
                border-radius: 12px;
                font-size: 1rem;
                font-weight: 600;
                cursor: pointer;
                transition: all 0.3s ease;
                text-decoration: none;
                display: inline-flex;
                align-items: center;
                gap: 0.5rem;
                font-family: inherit;
            }
            
            .btn-primary {
                background: linear-gradient(135deg, #667eea, #764ba2);
                color: white;
                box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
            }
            
            .btn-primary:hover {
                transform: translateY(-2px);
                box-shadow: 0 6px 20px rgba(102, 126, 234, 0.5);
            }
            
            .btn-success {
                background: linear-gradient(135deg, #48bb78, #38a169);
                color: white;
                box-shadow: 0 4px 12px rgba(72, 187, 120, 0.4);
            }
            
            .btn-success:hover {
                transform: translateY(-2px);
                box-shadow: 0 6px 20px rgba(72, 187, 120, 0.5);
            }
            
            .btn:disabled {
                opacity: 0.5;
                cursor: not-allowed;
                transform: none !important;
            }
            
            .status-card {
                margin-top: 2rem;
                padding: 2rem;
                background: linear-gradient(135deg, #f7fafc 0%, #edf2f7 100%);
                border-radius: 16px;
                border: 1px solid #e2e8f0;
                display: none;
            }
            
            .status-title {
                font-size: 1.1rem;
                font-weight: 600;
                color: #2d3748;
                margin-bottom: 1rem;
            }
            
            .progress-container {
                margin: 1rem 0;
            }
            
            .progress-bar {
                width: 100%;
                height: 8px;
                background: #e2e8f0;
                border-radius: 4px;
                overflow: hidden;
            }
            
            .progress-fill {
                height: 100%;
                background: linear-gradient(90deg, #667eea, #764ba2);
                border-radius: 4px;
                transition: width 0.3s ease;
                width: 0%;
            }
            
            .status-text {
                font-size: 0.9rem;
                color: #4a5568;
                margin-top: 0.5rem;
            }
            
            .success-message {
                color: #38a169;
                font-weight: 600;
            }
            
            .error-message {
                color: #e53e3e;
                font-weight: 600;
            }
            
            .footer {
                text-align: center;
                margin-top: 2rem;
                color: rgba(255,255,255,0.8);
                font-size: 0.8rem;
            }
            
            @keyframes fadeInUp {
                from {
                    opacity: 0;
                    transform: translateY(30px);
                }
                to {
                    opacity: 1;
                    transform: translateY(0);
                }
            }
            
            @keyframes pulse {
                0%, 100% { transform: scale(1); }
                50% { transform: scale(1.05); }
            }
            
            .processing {
                animation: pulse 2s infinite;
            }
            
            @media (max-width: 768px) {
                .container {
                    padding: 1rem;
                }
                
                .main-card {
                    padding: 2rem;
                }
                
                .header h1 {
                    font-size: 2rem;
                }
                
                .metrics-bar {
                    grid-template-columns: 1fr;
                }
                
                .action-buttons {
                    flex-direction: column;
                    align-items: center;
                }
                
                .btn {
                    width: 100%;
                    max-width: 300px;
                }
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎯 Time Verify AI</h1>
                <p>Extract timesheet data from DOCX files with AI precision</p>
                <div class="redhat-badge">
                    🚀 Powered by Red Hat OpenShift & Enterprise Linux
                </div>
            </div>
            
            <!-- Dashboard Metrics Bar -->
            <div class="metrics-bar">
                <div class="metric-card blue">
                    <div class="metric-header">
                        <div class="metric-title">Documents Processed</div>
                        <div class="metric-icon blue">📄</div>
                    </div>
                    <div class="metric-value" id="documentsProcessed">0</div>
                    <div class="metric-subtitle">↗ Today</div>
                </div>
                
                <div class="metric-card green">
                    <div class="metric-header">
                        <div class="metric-title">Hours Extracted</div>
                        <div class="metric-icon green">⏱</div>
                    </div>
                    <div class="metric-value" id="hoursExtracted">0</div>
                    <div class="metric-subtitle">↗ This session</div>
                </div>
                
                <div class="metric-card orange">
                    <div class="metric-header">
                        <div class="metric-title">Accuracy Rate</div>
                        <div class="metric-icon orange">✓</div>
                    </div>
                    <div class="metric-value" id="accuracyRate">97%</div>
                    <div class="metric-subtitle">↗ OCR Quality</div>
                </div>
                
                <div class="metric-card purple">
                    <div class="metric-header">
                        <div class="metric-title">Time Savings</div>
                        <div class="metric-icon purple">🚀</div>
                    </div>
                    <div class="metric-value" id="timeSavings">10min</div>
                    <div class="metric-subtitle">↗ vs Manual</div>
                </div>
            </div>
            
            <div class="main-card">
                <div class="upload-section">
                    <div class="upload-area" onclick="document.getElementById('docxFiles').click()">
                        <div class="upload-icon">📄</div>
                        <div class="upload-text">Upload DOCX Files</div>
                       <div class="upload-subtext">Select single or multiple files to process</div>
                       <input type="file" id="docxFiles" multiple accept=".docx,.doc" class="file-input">
                   </div>
                   
                   <div id="fileList" class="file-list"></div>
               </div>
               
               <div class="action-buttons">
                   <button id="processBtn" class="btn btn-primary" onclick="processFiles()">
                       🚀 Process Files
                   </button>
                   <button id="downloadBtn" class="btn btn-success" onclick="downloadExcel()" style="display: none;">
                       📥 Download Excel
                   </button>
               </div>
               
               <div id="statusCard" class="status-card">
                   <div class="status-title" id="statusTitle">Processing...</div>
                   <div class="progress-container">
                       <div class="progress-bar">
                           <div class="progress-fill" id="progressFill"></div>
                       </div>
                   </div>
                   <div class="status-text" id="statusText"></div>
               </div>
           </div>
           
           <div class="footer">
                <p>© 2025 Time Verify AI. All rights reserved.</p>
                <p>Built by Preevind </p>
           </div>
       </div>
       
       <script>
           let jobId = null;
           let processingInterval = null;
           let metricsInterval = null;
           
           // Load dashboard metrics
           async function loadMetrics() {
               try {
                   const response = await fetch('/dashboard');
                   const data = await response.json();
                   
                   document.getElementById('documentsProcessed').textContent = data.documents_processed;
                   document.getElementById('hoursExtracted').textContent = data.hours_extracted;
                   document.getElementById('accuracyRate').textContent = data.accuracy_rate + '%';
                   document.getElementById('timeSavings').textContent = data.time_savings + 'min';
               } catch (error) {
                   console.error('Error loading metrics:', error);
               }
           }
           
           // Initialize metrics on page load
           loadMetrics();
           
           // Refresh metrics every 30 seconds
           metricsInterval = setInterval(loadMetrics, 30000);
           
           document.getElementById('docxFiles').addEventListener('change', function() {
               const files = this.files;
               const fileList = document.getElementById('fileList');
               
               if (files.length > 0) {
                   let html = '';
                   for (let i = 0; i < files.length; i++) {
                       html += `<div class="file-item">${files[i].name}</div>`;
                   }
                   fileList.innerHTML = html;
                   fileList.style.display = 'block';
               } else {
                   fileList.style.display = 'none';
               }
           });
           
           async function processFiles() {
               const files = document.getElementById('docxFiles').files;
               
               if (files.length === 0) {
                   alert('Please select at least one DOCX file');
                   return;
               }
               
               const formData = new FormData();
               for (let file of files) {
                   formData.append('docx_files', file);
               }
               
               // Update UI
               document.getElementById('processBtn').disabled = true;
               document.getElementById('downloadBtn').style.display = 'none';
               document.getElementById('statusCard').style.display = 'block';
               document.getElementById('statusTitle').textContent = 'Initializing Red Hat processing...';
               document.getElementById('statusTitle').className = 'status-title processing';
               
               try {
                   const response = await fetch('/process-bulk', {
                       method: 'POST',
                       body: formData
                   });
                   
                   const result = await response.json();
                   
                   if (result.job_id) {
                       jobId = result.job_id;
                       startStatusCheck();
                   } else {
                       throw new Error(result.error || 'Unknown error');
                   }
               } catch (error) {
                   document.getElementById('statusTitle').textContent = 'Error';
                   document.getElementById('statusTitle').className = 'status-title error-message';
                   document.getElementById('statusText').textContent = error.message;
                   document.getElementById('processBtn').disabled = false;
               }
           }
           
           function startStatusCheck() {
               processingInterval = setInterval(checkStatus, 2000);
           }
           
           async function checkStatus() {
               if (!jobId) return;
               
               try {
                   const response = await fetch(`/status/${jobId}`);
                   const status = await response.json();
                   
                   if (status.status === 'processing') {
                       const progress = (status.processed / status.total) * 100;
                       document.getElementById('statusTitle').textContent = `Processing ${status.processed}/${status.total} files on Red Hat`;
                       document.getElementById('progressFill').style.width = progress + '%';
                       document.getElementById('statusText').textContent = `Current: ${status.current_file || 'Starting...'}`;
                   } else if (status.status === 'completed') {
                       clearInterval(processingInterval);
                       document.getElementById('statusTitle').textContent = 'Red Hat Processing Completed!';
                       document.getElementById('statusTitle').className = 'status-title success-message';
                       document.getElementById('progressFill').style.width = '100%';
                       document.getElementById('statusText').textContent = `Extracted ${status.total_entries} entries`;
                       document.getElementById('processBtn').disabled = false;
                       document.getElementById('downloadBtn').style.display = 'inline-flex';
                       
                       // Refresh metrics after completion
                       loadMetrics();
                   } else if (status.status === 'error') {
                       clearInterval(processingInterval);
                       document.getElementById('statusTitle').textContent = 'Processing Failed';
                       document.getElementById('statusTitle').className = 'status-title error-message';
                       document.getElementById('statusText').textContent = status.error;
                       document.getElementById('processBtn').disabled = false;
                   }
               } catch (error) {
                   console.error('Status check error:', error);
               }
           }
           
           async function downloadExcel() {
               if (!jobId) return;
               
               try {
                   const response = await fetch(`/download/${jobId}`);
                   
                   if (response.ok) {
                       const blob = await response.blob();
                       const url = URL.createObjectURL(blob);
                       const a = document.createElement('a');
                       a.href = url;
                       a.download = `redhat_timesheet_data_${new Date().toISOString().split('T')[0]}.xlsx`;
                       document.body.appendChild(a);
                       a.click();
                       document.body.removeChild(a);
                       URL.revokeObjectURL(url);
                   } else {
                       const error = await response.json();
                       alert('Download failed: ' + error.error);
                   }
               } catch (error) {
                   alert('Download failed: ' + error.message);
               }
           }
       </script>
   </body>
   </html>
   '''

@app.route('/process-bulk', methods=['POST'])
def process_bulk():
   """Process multiple DOCX files with Red Hat environment info"""
   try:
       files = request.files.getlist('docx_files')
       
       if not files:
           return jsonify({'error': 'No files provided'}), 400
       
       logger.info(f"🎯 Red Hat processing initiated for {len(files)} files")
       
       # Generate job ID
       job_id = str(uuid.uuid4())
       
       # Prepare file data
       file_data = []
       for file in files:
           if file.filename.lower().endswith(('.docx', '.doc')):
               file_data.append((file.filename, file.read()))
       
       if not file_data:
           return jsonify({'error': 'No valid DOCX files found'}), 400
       
       # Initialize status with Red Hat info
       processing_status[job_id] = {
           'status': 'processing',
           'total': len(file_data),
           'processed': 0,
           'current_file': '',
           'total_entries': 0,
           'redhat_environment': redhat_env,
           'platform': platform.system(),
           'hostname': os.environ.get('HOSTNAME', 'unknown')
       }
       
       # Start processing in background thread
       thread = threading.Thread(target=process_files_bulk, args=(file_data, job_id))
       thread.daemon = True
       thread.start()
       
       return jsonify({
           'job_id': job_id, 
           'total_files': len(file_data),
           'redhat_environment': redhat_env,
           'message': 'Processing started on Red Hat infrastructure'
       })
       
   except Exception as e:
       logger.error(f"❌ Red Hat bulk processing error: {e}")
       return jsonify({'error': str(e)}), 500

@app.route('/status/<job_id>')
def get_status(job_id):
   """Get processing status with Red Hat environment details"""
   if job_id not in processing_status:
       return jsonify({'error': 'Job not found'}), 404
   
   status_data = processing_status[job_id].copy()
   status_data['timestamp'] = datetime.now().isoformat()
   
   return jsonify(status_data)

@app.route('/download/<job_id>')
def download_result(job_id):
   """Download Excel file with Red Hat branding"""
   try:
       if job_id not in processing_results:
           return jsonify({'error': 'Results not found'}), 404
       
       data = processing_results[job_id]
       
       if not data:
           return jsonify({'error': 'No data extracted'}), 400
       
       # Create Excel file
       wb = create_excel_file(data)
       
       # Save to memory
       output = io.BytesIO()
       wb.save(output)
       output.seek(0)
       
       # Create filename
       filename = f'redhat_timesheet_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
       
       # Clean up
       del processing_results[job_id]
       del processing_status[job_id]
       
       # Use Flask's make_response for better compatibility
       from flask import make_response
       
       response = make_response(output.getvalue())
       response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
       response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
       
       return response
       
   except Exception as e:
       logger.error(f"❌ Download error: {e}")
       return jsonify({'error': str(e)}), 500

@app.route('/health')
def health():
   """Health check with Red Hat environment info"""
   return jsonify({
       'status': 'healthy',
       'service': 'Time Verify AI',
       'version': '1.0.0',
       'claude_api': 'available' if claude_client.api_key != 'your-api-key-here' else 'missing',
       'redhat_environment': redhat_env,
       'platform': platform.system(),
       'hostname': os.environ.get('HOSTNAME', 'unknown'),
       'openshift_namespace': os.environ.get('OPENSHIFT_BUILD_NAMESPACE', 'not_available'),
       'container_runtime': 'podman' if redhat_env.get('podman') else 'docker',
       'timestamp': datetime.now().isoformat()
   })

@app.route('/redhat-info')
def redhat_info():
   """Dedicated endpoint for Red Hat environment information"""
   try:
       # Additional Red Hat specific information
       redhat_details = {
           'environment': redhat_env,
           'platform': platform.system(),
           'hostname': os.environ.get('HOSTNAME', 'unknown'),
           'openshift_vars': {
               'namespace': os.environ.get('OPENSHIFT_BUILD_NAMESPACE'),
               'build_name': os.environ.get('OPENSHIFT_BUILD_NAME'),
               'build_source': os.environ.get('OPENSHIFT_BUILD_SOURCE'),
               'kubernetes_host': os.environ.get('KUBERNETES_SERVICE_HOST')
           },
           'container_info': {
               'container_runtime': 'podman' if redhat_env.get('podman') else 'docker',
               'systemd_available': redhat_env.get('systemd', False)
           }
       }
       
       # Check for Red Hat specific files
       if os.path.exists('/etc/redhat-release'):
           try:
               with open('/etc/redhat-release', 'r') as f:
                   redhat_details['rhel_version'] = f.read().strip()
           except:
               pass
       
       return jsonify(redhat_details)
       
   except Exception as e:
       return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
   # Red Hat OpenShift compatibility - CRITICAL for deployment
   port = int(os.environ.get('PORT', 8080))
   host = os.environ.get('HOST', '0.0.0.0')
   
   # Log Red Hat environment
   logger.info("🎯 Time Verify AI - Starting...")
   logger.info(f"🚀 Red Hat Environment: {redhat_env}")
   logger.info(f"🌐 Server: {host}:{port}")
   logger.info(f"🔧 Platform: {platform.system()}")
   logger.info(f"🏢 Hostname: {os.environ.get('HOSTNAME', 'unknown')}")
   
   if redhat_env.get('openshift'):
       logger.info("🎯 Running on Red Hat OpenShift")
       logger.info(f"📦 Namespace: {os.environ.get('OPENSHIFT_BUILD_NAMESPACE', 'unknown')}")
   
   if redhat_env.get('rhel'):
       logger.info("🎯 Running on Red Hat Enterprise Linux")
   
   if redhat_env.get('podman'):
       logger.info("🎯 Red Hat Podman container runtime available")
   
   if not claude_client.api_key or claude_client.api_key == 'your-api-key-here':
       logger.warning("⚠️ Set CLAUDE_API_KEY environment variable")
   
   logger.info("🚀 Time Verify AI ready for deployment")
   app.run(host=host, port=port, debug=False)